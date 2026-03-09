"""Tests for trending output across all 6 formats."""

import csv
import io
import json
from unittest.mock import MagicMock

import pytest

from cja_auto_sdr.org.models import (
    ComponentDistribution,
    ComponentInfo,
    DataViewSummary,
    OrgReportConfig,
    OrgReportResult,
    OrgReportTrending,
    TrendingDelta,
    TrendingSnapshot,
)
from cja_auto_sdr.org.writers import (
    build_org_report_json_data,
    write_org_report_console,
    write_org_report_csv,
    write_org_report_excel,
    write_org_report_html,
    write_org_report_markdown,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_trending():
    """Build a minimal OrgReportTrending for testing."""
    return OrgReportTrending(
        snapshots=[
            TrendingSnapshot(
                timestamp="2026-01-01T00:00:00Z",
                data_view_count=10,
                component_count=100,
                core_count=80,
                isolated_count=20,
                high_sim_pair_count=2,
                dv_ids={"dv1", "dv2"},
            ),
            TrendingSnapshot(
                timestamp="2026-02-01T00:00:00Z",
                data_view_count=12,
                component_count=120,
                core_count=95,
                isolated_count=25,
                high_sim_pair_count=3,
                dv_ids={"dv1", "dv2", "dv3"},
            ),
        ],
        deltas=[
            TrendingDelta(
                from_timestamp="2026-01-01T00:00:00Z",
                to_timestamp="2026-02-01T00:00:00Z",
                data_view_delta=2,
                component_delta=20,
                core_delta=15,
                isolated_delta=5,
                high_sim_pair_delta=1,
            ),
        ],
        drift_scores={"dv1": 0.82, "dv2": 0.15, "dv3": 0.45},
        window_size=2,
    )


def _make_result():
    """Build a minimal OrgReportResult."""
    return OrgReportResult(
        timestamp="2026-02-01T00:00:00Z",
        org_id="test_org",
        parameters=OrgReportConfig(),
        data_view_summaries=[
            DataViewSummary(data_view_id="dv1", data_view_name="Test DV 1", metric_count=50, dimension_count=30),
            DataViewSummary(data_view_id="dv2", data_view_name="Test DV 2", metric_count=20, dimension_count=10),
        ],
        component_index={"m1": ComponentInfo(component_id="m1", component_type="metric", data_views={"dv1", "dv2"})},
        distribution=ComponentDistribution(
            core_metrics=["m1"],
            isolated_metrics=[],
            core_dimensions=[],
            isolated_dimensions=[],
        ),
        similarity_pairs=None,
        recommendations=[],
        duration=1.5,
    )


# ---------------------------------------------------------------------------
# Console
# ---------------------------------------------------------------------------


class TestConsoleWithTrending:
    def test_trending_none_unchanged(self, capsys):
        result = _make_result()
        config = OrgReportConfig()
        write_org_report_console(result, config, quiet=False, trending=None)
        output = capsys.readouterr().out
        assert "TRENDING" not in output

    def test_trending_renders_section(self, capsys):
        result = _make_result()
        config = OrgReportConfig()
        trending = _make_trending()
        write_org_report_console(result, config, quiet=False, trending=trending)
        output = capsys.readouterr().out
        assert "TRENDING" in output
        assert "Data Views" in output
        assert "Components" in output
        assert "Top Drift" in output
        assert "dv1" in output

    def test_trending_quiet_suppressed(self, capsys):
        result = _make_result()
        config = OrgReportConfig()
        write_org_report_console(result, config, quiet=True, trending=_make_trending())
        output = capsys.readouterr().out
        assert "TRENDING" not in output


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------


class TestJsonWithTrending:
    def test_trending_none_no_key(self):
        result = _make_result()
        data = build_org_report_json_data(result, trending=None)
        assert "trending" not in data

    def test_trending_present_adds_key(self):
        result = _make_result()
        trending = _make_trending()
        data = build_org_report_json_data(result, trending=trending)
        assert "trending" in data
        t = data["trending"]
        assert "snapshots" in t
        assert "deltas" in t
        assert "drift_scores" in t
        assert len(t["snapshots"]) == 2
        assert len(t["deltas"]) == 1
        assert t["drift_scores"]["dv1"] == 0.82

    def test_snapshot_fields(self):
        result = _make_result()
        trending = _make_trending()
        data = build_org_report_json_data(result, trending=trending)
        snap = data["trending"]["snapshots"][0]
        assert snap["timestamp"] == "2026-01-01T00:00:00Z"
        assert snap["data_view_count"] == 10
        assert snap["component_count"] == 100
        assert snap["core_count"] == 80
        assert snap["isolated_count"] == 20
        assert snap["high_sim_pair_count"] == 2

    def test_delta_fields(self):
        result = _make_result()
        trending = _make_trending()
        data = build_org_report_json_data(result, trending=trending)
        delta = data["trending"]["deltas"][0]
        assert delta["data_view_delta"] == 2
        assert delta["component_delta"] == 20

    def test_json_roundtrip(self):
        """Trending JSON is serializable and roundtrippable."""
        result = _make_result()
        trending = _make_trending()
        data = build_org_report_json_data(result, trending=trending)
        serialized = json.dumps(data, default=str)
        parsed = json.loads(serialized)
        assert parsed["trending"]["window_size"] == 2


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


class TestExcelWithTrending:
    def test_trending_none_no_worksheet(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        path = write_org_report_excel(result, tmp_path / "test.xlsx", str(tmp_path), logger, trending=None)
        import openpyxl

        wb = openpyxl.load_workbook(path)
        assert "Trending" not in wb.sheetnames
        wb.close()

    def test_trending_adds_worksheet(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        trending = _make_trending()
        path = write_org_report_excel(result, tmp_path / "test.xlsx", str(tmp_path), logger, trending=trending)
        import openpyxl

        wb = openpyxl.load_workbook(path)
        assert "Trending" in wb.sheetnames
        ws = wb["Trending"]
        # Should have header row + metric rows
        assert ws.max_row >= 2
        wb.close()


# ---------------------------------------------------------------------------
# Markdown
# ---------------------------------------------------------------------------


class TestMarkdownWithTrending:
    def test_trending_none_unchanged(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        path = write_org_report_markdown(result, tmp_path / "test.md", str(tmp_path), logger, trending=None)
        content = open(path, encoding="utf-8").read()
        assert "Trending" not in content

    def test_trending_adds_section(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        trending = _make_trending()
        path = write_org_report_markdown(result, tmp_path / "test.md", str(tmp_path), logger, trending=trending)
        content = open(path, encoding="utf-8").read()
        assert "## Trending" in content
        assert "Data Views" in content
        assert "Drift Scores" in content


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------


class TestHtmlWithTrending:
    def test_trending_none_unchanged(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        path = write_org_report_html(result, tmp_path / "test.html", str(tmp_path), logger, trending=None)
        content = open(path, encoding="utf-8").read()
        assert "trending" not in content.lower() or "Trending" not in content

    def test_trending_adds_section(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        trending = _make_trending()
        path = write_org_report_html(result, tmp_path / "test.html", str(tmp_path), logger, trending=trending)
        content = open(path, encoding="utf-8").read()
        assert "Trending" in content
        assert "drift" in content.lower()


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


class TestCsvWithTrending:
    def test_trending_none_no_trending_file(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        write_org_report_csv(result, None, str(tmp_path), logger, trending=None)
        trending_files = list(tmp_path.glob("**/*trending*"))
        assert trending_files == []

    def test_trending_creates_files(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        trending = _make_trending()
        write_org_report_csv(result, None, str(tmp_path), logger, trending=trending)
        trending_files = list(tmp_path.glob("**/*trending*"))
        assert len(trending_files) >= 1
