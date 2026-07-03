"""Characterization tests for the diff Excel writer's color-coded rows.

Pins exact data-cell values and per-row fill colors for
``write_diff_excel_output`` (component-diff sheets: Metrics/Dimensions Diff)
and its inventory counterpart (Calc Metrics/Segments Diff). These values were
captured from the pre-refactor implementation (per-cell ``df.iloc`` scalar
writes) and must remain byte-identical after the ``write_row`` refactor.
"""

from __future__ import annotations

import logging

from openpyxl import load_workbook

from cja_auto_sdr.diff.models import (
    ChangeType,
    ComponentDiff,
    DiffResult,
    DiffSummary,
    InventoryItemDiff,
    MetadataDiff,
)
from cja_auto_sdr.output.diff.excel import write_diff_excel_output

logger = logging.getLogger(__name__)

# Fill colors from the format definitions in write_diff_excel_output:
#   added_format    -> bg_color #d4edda
#   removed_format  -> bg_color #f8d7da
#   modified_format -> bg_color #fff3cd
#   normal_format   -> no bg_color (unfilled)
ADDED_RGB = "FFD4EDDA"
REMOVED_RGB = "FFF8D7DA"
MODIFIED_RGB = "FFFFF3CD"
UNCHANGED_RGB = "00000000"  # xlsxwriter default: no fill applied


def _build_diff_result() -> DiffResult:
    """Build a DiffResult exercising ADDED / REMOVED / MODIFIED / UNCHANGED
    for both component diffs (Metrics) and inventory diffs (Calc Metrics)."""
    metric_diffs = [
        ComponentDiff(
            id="m_added",
            name="Added Metric",
            change_type=ChangeType.ADDED,
            source_data=None,
            target_data={"name": "Added Metric"},
        ),
        ComponentDiff(
            id="m_removed",
            name="Removed Metric",
            change_type=ChangeType.REMOVED,
            source_data={"name": "Removed Metric"},
            target_data=None,
        ),
        ComponentDiff(
            id="m_modified",
            name="Modified Metric",
            change_type=ChangeType.MODIFIED,
            source_data={"description": "old desc"},
            target_data={"description": "new desc"},
            changed_fields={"description": ("old desc", "new desc")},
        ),
        ComponentDiff(
            id="m_unchanged",
            name="Unchanged Metric",
            change_type=ChangeType.UNCHANGED,
            source_data={"description": "same"},
            target_data={"description": "same"},
        ),
    ]

    dimension_diffs = [
        ComponentDiff(id="d_unchanged", name="Unchanged Dimension", change_type=ChangeType.UNCHANGED),
    ]

    calc_metrics_diffs = [
        InventoryItemDiff(
            id="cm_added", name="Added CM", change_type=ChangeType.ADDED, inventory_type="calculated_metric"
        ),
        InventoryItemDiff(
            id="cm_removed", name="Removed CM", change_type=ChangeType.REMOVED, inventory_type="calculated_metric"
        ),
        InventoryItemDiff(
            id="cm_modified",
            name="Modified CM",
            change_type=ChangeType.MODIFIED,
            inventory_type="calculated_metric",
            changed_fields={"formula": ("old formula", "new formula")},
        ),
        InventoryItemDiff(
            id="cm_unchanged", name="Unchanged CM", change_type=ChangeType.UNCHANGED, inventory_type="calculated_metric"
        ),
    ]

    segments_diffs = [
        InventoryItemDiff(id="s_added", name="Added Segment", change_type=ChangeType.ADDED, inventory_type="segment"),
        InventoryItemDiff(
            id="s_removed", name="Removed Segment", change_type=ChangeType.REMOVED, inventory_type="segment"
        ),
        InventoryItemDiff(
            id="s_modified",
            name="Modified Segment",
            change_type=ChangeType.MODIFIED,
            inventory_type="segment",
            changed_fields={"definition": ("old def", "new def")},
        ),
        InventoryItemDiff(
            id="s_unchanged", name="Unchanged Segment", change_type=ChangeType.UNCHANGED, inventory_type="segment"
        ),
    ]

    summary = DiffSummary(
        source_metrics_count=4,
        target_metrics_count=4,
        metrics_added=1,
        metrics_removed=1,
        metrics_modified=1,
        metrics_unchanged=1,
        source_dimensions_count=1,
        target_dimensions_count=1,
        dimensions_unchanged=1,
        source_calc_metrics_count=4,
        target_calc_metrics_count=4,
        calc_metrics_added=1,
        calc_metrics_removed=1,
        calc_metrics_modified=1,
        calc_metrics_unchanged=1,
        source_segments_count=4,
        target_segments_count=4,
        segments_added=1,
        segments_removed=1,
        segments_modified=1,
        segments_unchanged=1,
    )

    metadata = MetadataDiff(
        source_name="Source DV",
        target_name="Target DV",
        source_id="dv_source",
        target_id="dv_target",
    )

    return DiffResult(
        summary=summary,
        metadata_diff=metadata,
        metric_diffs=metric_diffs,
        dimension_diffs=dimension_diffs,
        calc_metrics_diffs=calc_metrics_diffs,
        segments_diffs=segments_diffs,
    )


def test_diff_excel_cells_and_fills_unchanged(temp_output_dir):
    """Metrics Diff sheet: cell values + per-row fill colors are pinned."""
    diff_result = _build_diff_result()

    filepath = write_diff_excel_output(diff_result, "test_diff", temp_output_dir, logger)

    wb = load_workbook(filepath)
    ws = wb["Metrics Diff"]

    expected_statuses = ["ADDED", "REMOVED", "MODIFIED", "UNCHANGED"]
    statuses = [ws.cell(row=r, column=1).value for r in range(2, 2 + len(expected_statuses))]
    assert statuses == expected_statuses

    expected_rows = [
        ("ADDED", "m_added", "Added Metric", None),
        ("REMOVED", "m_removed", "Removed Metric", None),
        ("MODIFIED", "m_modified", "Modified Metric", "description: 'old desc' -> 'new desc'"),
        ("UNCHANGED", "m_unchanged", "Unchanged Metric", None),
    ]
    for offset, expected_row in enumerate(expected_rows):
        row_idx = 2 + offset
        actual_row = tuple(ws.cell(row=row_idx, column=c).value for c in range(1, 5))
        assert actual_row == expected_row

    expected_fills = [ADDED_RGB, REMOVED_RGB, MODIFIED_RGB, UNCHANGED_RGB]
    for offset, expected_rgb in enumerate(expected_fills):
        row_idx = 2 + offset
        fill = ws.cell(row=row_idx, column=1).fill
        assert fill.fgColor.rgb == expected_rgb

    # Header row untouched by the color pass (still whatever to_excel wrote).
    assert [ws.cell(row=1, column=c).value for c in range(1, 5)] == ["Status", "ID", "Name", "Details"]


def test_diff_excel_dimensions_sheet_unchanged_row(temp_output_dir):
    """Dimensions Diff sheet: single UNCHANGED row has no fill."""
    diff_result = _build_diff_result()

    filepath = write_diff_excel_output(diff_result, "test_diff", temp_output_dir, logger)

    wb = load_workbook(filepath)
    ws = wb["Dimensions Diff"]

    assert ws.cell(row=2, column=1).value == "UNCHANGED"
    assert tuple(ws.cell(row=2, column=c).value for c in range(1, 5)) == (
        "UNCHANGED",
        "d_unchanged",
        "Unchanged Dimension",
        None,
    )
    assert ws.cell(row=2, column=1).fill.fgColor.rgb == UNCHANGED_RGB


def test_diff_excel_inventory_sheets_cells_and_fills_unchanged(temp_output_dir):
    """write_inventory_diff_sheet gets the identical transform: pin Calc Metrics + Segments sheets."""
    diff_result = _build_diff_result()

    filepath = write_diff_excel_output(diff_result, "test_diff", temp_output_dir, logger)

    wb = load_workbook(filepath)

    expected_rows_by_sheet = {
        "Calc Metrics Diff": [
            ("ADDED", "cm_added", "Added CM", None),
            ("REMOVED", "cm_removed", "Removed CM", None),
            ("MODIFIED", "cm_modified", "Modified CM", "formula: 'old formula' -> 'new formula'"),
            ("UNCHANGED", "cm_unchanged", "Unchanged CM", None),
        ],
        "Segments Diff": [
            ("ADDED", "s_added", "Added Segment", None),
            ("REMOVED", "s_removed", "Removed Segment", None),
            ("MODIFIED", "s_modified", "Modified Segment", "definition: 'old def' -> 'new def'"),
            ("UNCHANGED", "s_unchanged", "Unchanged Segment", None),
        ],
    }
    expected_fills = [ADDED_RGB, REMOVED_RGB, MODIFIED_RGB, UNCHANGED_RGB]

    for sheet_name, expected_rows in expected_rows_by_sheet.items():
        ws = wb[sheet_name]
        for offset, (expected_row, expected_rgb) in enumerate(zip(expected_rows, expected_fills, strict=True)):
            row_idx = 2 + offset
            actual_row = tuple(ws.cell(row=row_idx, column=c).value for c in range(1, 5))
            assert actual_row == expected_row, f"{sheet_name} row {row_idx}"
            fill = ws.cell(row=row_idx, column=1).fill
            assert fill.fgColor.rgb == expected_rgb, f"{sheet_name} row {row_idx} fill"
